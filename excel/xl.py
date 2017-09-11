import openpyxl

class xll:
    feldlst=list()
    queslst=list()
    typlst=list()

    def feld (self,num):
        x=list()
        for i in range (1,self.ws.max_column):
            x.append(self.ws[num][i].value)
        return x

    def __init__(self,file1,sheet1):
        self.wb=openpyxl.load_workbook(file1)
        self.ws=self.wb.get_sheet_by_name(sheet1)
        self.feldlst=self.feld(1)
        self.queslst=self.feld(2)
        self.typlst=self.feld(3)
